"""
Module 5: 钻孔电视语义分割 — Excel 裂隙参数 → 正弦曲线 mask → U-Net 训练/推理
"""
import json
import math
import os
import re

import cv2
import numpy as np

# ---- 物理参数 ----
BOREHOLE_DIAMETER_MM = 75.0  # CK12 为 75mm，其他钻孔可能不同


def _imread_rgb(path):
    """用 PIL 读取图像 → numpy RGB 数组 → 转为 BGR（兼容 OpenCV 和中文路径）"""
    from PIL import Image
    img_pil = Image.open(path).convert("RGB")
    img_np = np.array(img_pil)
    return cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)


# ================================================================
# 1. Excel 参数解析
# ================================================================

def read_fractures_from_excel(excel_path):
    """
    读取成果表 Excel，返回每条裂隙的参数字典。
    字段映射（CK12）：
      C3 (col 3) = 倾角 (°) = dip_angle
      C4 (col 4) = 倾向 (°) = dip_direction
      C5 (col 5) = 顶点深度 (m) = z0 (the depth at the crest of the sine)
      C6 (col 6) = 顶点高度 (m) = elevation at crest (not used for mask)
    使用 openpyxl 读取，跳过表头行。
    返回: [{"z0": float, "dip": float, "phi": float}, ...]
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    fractures = []
    for row in range(7, ws.max_row + 1):
        dip = ws.cell(row, 3).value
        direction = ws.cell(row, 4).value
        z0 = ws.cell(row, 5).value
        # 跳过空行和合并单元格描述行
        if dip is None or not isinstance(dip, (int, float)):
            continue
        if z0 is None or not isinstance(z0, (int, float)):
            continue

        fractures.append({
            "z0": float(z0),
            "dip": float(dip),            # 倾角 (dip angle, °)
            "phi": float(direction if direction else 0),  # 倾向 (dip direction, °)
        })
    return fractures


# ---- 文件名深度解析 ----
def parse_tv_filename(filename):
    """
    从电视图像文件名提取深度区间。
    例: "12_762_95_0_20250708_153813_512_000.jpg"
      → interval_mm = 512/1000 = 0.512m
      → start_mm = 0/1000 = 0m
      → z_top = 0.0, z_bottom = 0.512
    """
    # 提取末尾两个数字: _512_000.jpg
    m = re.search(r'_(\d+)_(\d+)\.(jpg|JPG|png|PNG)$', filename)
    if m:
        interval_mm = int(m.group(1))
        start_mm = int(m.group(2))
        z_top = start_mm / 1000.0
        z_bottom = z_top + interval_mm / 1000.0
        return z_top, z_bottom
    return None, None


# ================================================================
# 2. 正弦曲线绘制 → mask
# ================================================================

def fracture_to_sine_pixels(z0, dip, phi, z_top, z_bottom,
                            img_w, img_h, borehole_diameter_mm=None):
    """
    将一条裂隙参数转换为图像上的正弦曲线像素点。
    返回 [(x, y), ...] 有效点列表。
    """
    if borehole_diameter_mm is None:
        borehole_diameter_mm = BOREHOLE_DIAMETER_MM

    r_m = borehole_diameter_mm / 2000.0  # mm → m
    dip_rad = math.radians(dip)
    phi_rad = math.radians(phi)

    # 振幅 (m): A = r * cot(dip)
    A_m = r_m / math.tan(dip_rad) if math.tan(dip_rad) > 0.0001 else 0

    depth_range = z_bottom - z_top
    if depth_range <= 0:
        return []

    points = []
    # 图像实际有效区域：左边 93px 为刻度尺，岩壁图像从 X_LEFT 到 X_RIGHT
    X_LEFT = 93
    X_RIGHT = 575
    borehole_w = X_RIGHT - X_LEFT
    for theta in range(0, 361, 1):  # 1° steps
        theta_rad = math.radians(theta)
        z = z0 + A_m * math.sin(theta_rad - phi_rad)
        x_px = X_LEFT + int(theta / 360.0 * borehole_w)
        y_px = int((z - z_top) / depth_range * img_h)
        if 0 <= y_px < img_h:
            points.append((x_px, y_px))
    return points


def generate_mask_for_image(image_path, fractures, output_path=None,
                             borehole_diameter_mm=None, line_width=5):
    """
    为一张电视图像生成裂隙 mask。
    读取图像尺寸，自动匹配深度区间内的裂隙，绘制正弦曲线。
    """
    img = _imread_rgb(image_path)
    h, w = img.shape[:2]

    z_top, z_bottom = parse_tv_filename(os.path.basename(image_path))
    if z_top is None:
        raise ValueError(f"无法从文件名解析深度: {image_path}")

    mask = np.zeros((h, w), dtype=np.uint8)
    matched = 0

    for fr in fractures:
        z0 = fr["z0"]
        # 裂隙顶点深度是否在当前图像范围内
        if not (z_top <= z0 <= z_bottom):
            continue

        pts = fracture_to_sine_pixels(
            z0, fr["dip"], fr["phi"],
            z_top, z_bottom, w, h, borehole_diameter_mm
        )
        if len(pts) >= 2:
            pts_arr = np.array(pts, dtype=np.int32)
            cv2.polylines(mask, [pts_arr], isClosed=False,
                          color=255, thickness=line_width)
            matched += 1

    if output_path:
        cv2.imwrite(output_path, mask)
    return mask, matched


def generate_overlay(image_path, mask, output_path, alpha=0.5):
    """将 mask 半透明叠加到原图上，用于人工检查"""
    img = _imread_rgb(image_path)
    overlay = img.copy()
    overlay[mask > 0] = (0, 0, 255)  # 红色
    blended = cv2.addWeighted(img, 1 - alpha, overlay, alpha, 0)

    if output_path:
        cv2.imwrite(output_path, blended)
    return blended


# ================================================================
# 3. 批量生成
# ================================================================

def generate_all_masks(tv_image_dir, fractures, output_dir,
                        borehole_diameter_mm=None, line_width=5,
                        progress_callback=None):
    """批量生成 TV 图像目录下的所有 mask + 叠加图"""
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, "overlays"), exist_ok=True)

    jpgs = sorted([f for f in os.listdir(tv_image_dir)
                   if f.lower().endswith(('.jpg', '.jpeg', '.png'))])

    done = 0
    for i, fname in enumerate(jpgs):
        path = os.path.join(tv_image_dir, fname)
        mask_path = os.path.join(output_dir, fname.replace(".jpg", "_mask.png"))
        overlay_path = os.path.join(output_dir, "overlays", fname)

        try:
            mask, n = generate_mask_for_image(
                path, fractures, mask_path, borehole_diameter_mm, line_width
            )
            generate_overlay(path, mask, overlay_path)
            done += 1
        except Exception as e:
            print(f"  Error {fname}: {e}")

        if progress_callback:
            progress_callback(i + 1, len(jpgs))

    return done


# ================================================================
# 4. U-Net 训练（预留接口）
# ================================================================

def train_unet(data_yaml, epochs=50, device="cuda"):
    """预留：训练 U-Net 分割模型"""
    raise NotImplementedError(
        "U-Net 训练将在 mask 验证通过后实现。"
    )


# ================================================================
# 5. U-Net 推理（预留接口）
# ================================================================

def predict_unet(model_path, image_path, output_path=None):
    """预留：U-Net 推理"""
    raise NotImplementedError(
        "U-Net 推理将在模型训练完成后实现。"
    )
